import sys
import leveldb
from openpyxl import load_workbook

if len(sys.argv) < 3:
    print("Usage: python3 excel2leveldb.py <leveldb_path> <excel_file_path>")
    sys.exit(1)

leveldb_path = sys.argv[1]
excel_file_path = sys.argv[2]

# Load the Excel file
workbook = load_workbook(filename=excel_file_path)
worksheet = workbook.active

# Create a new LevelDB database
db = leveldb.LevelDB(leveldb_path)

# Iterate over the rows and columns of the Excel file
for row in worksheet.iter_rows(min_row=1, values_only=True):
    key = row[0]
    value = row[1]

    # Convert the key and value to bytes if necessary
    if not isinstance(key, bytes):
        key = str(key).encode('utf-8')
    if not isinstance(value, bytes):
        value = str(value).encode('utf-8')

    # Insert the key-value pair into the LevelDB database
    db.Put(key, value)

# Close the LevelDB database
db = None
